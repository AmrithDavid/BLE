import asyncio
import threading
from bleak import BleakClient
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import tkinter as tk
from tkinter import ttk
from queue import Queue
import time
import struct
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import datetime
import configparser
import sys
import os

if os.path.exists('FSMApp.ini'):
    cfg = configparser.ConfigParser()
    cfg.read('FSMApp.ini')
else:    
    print("Settings not found")
    time.sleep(1)
    sys.exit()    


# Load settings
try:
    DEVICE_ADDRESS = cfg['SETTINGS']['Device_address']
    FSM_DATA_CHAR_UUID = cfg['SETTINGS']['FSM_data_characteristic_UUID']
    PLOTSTYLE = cfg['SETTINGS']['Plot_style']
except KeyError:
    print("Settings not found")
    time.sleep(1)
    sys.exit()

# Global variables for data
time_data = []
sample_time_data = []
array_a_data = [[] for _ in range(21)]  # 6 wavelengths * 3 PDs + 1 LED OFF * 3 PDs
array_b_data = [[] for _ in range(21)]  # 6 wavelengths * 3 PDs + 1 LED OFF * 3 PDs
data_queue = Queue()
start_time = None
running = True
start_timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
filename_csv = f"fsm_data_{start_timestamp}.csv"
 
# Wavelength options
wavelength_options = ["784 nm", "800 nm", "818 nm", "835 nm", "851 nm", "881 nm", "LED OFF"]
 
def parse_fsm_data(data):
    # Parse time data (bytes 1-6)
    time_min, time_us = struct.unpack_from('<HI', data, 0)
    total_time = time_min * 60 + time_us / 1_000_000
 
    # Function to parse PD values (28 unsigned integers)
    def parse_array_data(start_index):
        values = struct.unpack_from('<28I', data, start_index)
        return np.array(values) / 1e9  # Convert to volts
 
    # Parse Array A data (bytes 9-93)
    array_a_data = parse_array_data(9)
 
    # Parse Array B data (bytes 96-180)
    array_b_data = parse_array_data(96)
 
    return total_time, array_a_data, array_b_data
 
def notification_handler(sender, data):
    global start_time
    if start_time is None:
        start_time = time.time()
   
    try:
        sample_time, array_a_data, array_b_data = parse_fsm_data(data)
        timestamp = time.time() - start_time
        data_queue.put((timestamp, sample_time, array_a_data, array_b_data))
        csv_update(timestamp, sample_time, array_a_data, array_b_data)
        print(f"Time: {timestamp:.2f}s, Sample Time: {sample_time:.2f}s, Data received and queued")
    except Exception as e:
        print(f"Error parsing data: {e}")
 
def create_gui():
    root = tk.Tk()
    root.title("FSM Data Visualization")
    root.geometry("800x600")  # Set initial window size
 
    # Make the window resizable
    root.grid_rowconfigure(0, weight=1)
    root.grid_columnconfigure(0, weight=1)
 
    main_frame = ttk.Frame(root)
    main_frame.grid(row=0, column=0, sticky="nsew")
    main_frame.grid_rowconfigure(0, weight=1)
    main_frame.grid_columnconfigure(0, weight=1)
 
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(8, 8))
    canvas = FigureCanvasTkAgg(fig, master=main_frame)
    canvas_widget = canvas.get_tk_widget()
    canvas_widget.grid(row=0, column=0, sticky="nsew")
 
    ax1.set_title("LED A")
    ax2.set_title("LED B")
 
    control_frame = ttk.Frame(main_frame)
    control_frame.grid(row=1, column=0, sticky="ew", padx=5, pady=5)
    control_frame.grid_columnconfigure(1, weight=1)
    control_frame.grid_columnconfigure(3, weight=1)
 
    led_a_var = tk.StringVar(value=wavelength_options[0])
    led_b_var = tk.StringVar(value=wavelength_options[0])
 
    ttk.Label(control_frame, text="LED A:").grid(row=0, column=0, sticky="e")
    led_a_dropdown = ttk.Combobox(control_frame, textvariable=led_a_var, values=wavelength_options, width=10)
    led_a_dropdown.grid(row=0, column=1, sticky="ew", padx=(0, 10))
 
    ttk.Label(control_frame, text="LED B:").grid(row=0, column=2, sticky="e")
    led_b_dropdown = ttk.Combobox(control_frame, textvariable=led_b_var, values=wavelength_options, width=10)
    led_b_dropdown.grid(row=0, column=3, sticky="ew")
 
    lines_a = [ax1.plot([], [], label=f'PD{i+1}')[0] for i in range(3)]
    lines_b = [ax2.plot([], [], label=f'PD{i+1}')[0] for i in range(3)]
 
    for ax in (ax1, ax2):
        ax.legend(loc='lower right')
        ax.set_xlabel("Time (s)")
        ax.set_ylabel("Voltage (V)")
        ax.grid(True)
        ax.set_yscale(PLOTSTYLE)
 
    def update_plot():
        if time_data:
            led_a_index = wavelength_options.index(led_a_var.get()) * 3
            led_b_index = wavelength_options.index(led_b_var.get()) * 3
 
            for i in range(3):
                lines_a[i].set_data(time_data, array_a_data[led_a_index + i])
                lines_b[i].set_data(time_data, array_b_data[led_b_index + i])
 
            for ax in (ax1, ax2):
                ax.relim()
                ax.autoscale_view()
 
            canvas.draw_idle()
 
    def process_queue():
        while not data_queue.empty():
            timestamp, sample_time, a_data, b_data = data_queue.get()
            time_data.append(timestamp)
            sample_time_data.append(sample_time)
            for i in range(21):
                array_a_data[i].append(a_data[i])
                array_b_data[i].append(b_data[i])
       
        update_plot()
       
        if running:
            root.after(100, process_queue)
 
    def on_dropdown_change(*args):
        root.after(10, update_plot)
 
    def on_closing():
        global running
        running = False
        root.quit()
 
    def on_resize(event):
        # Update the figure size when the window is resized
        fig.set_size_inches(event.width/100, event.height/100)
        fig.tight_layout()
        canvas.draw_idle()
 
    root.protocol("WM_DELETE_WINDOW", on_closing)
    led_a_var.trace_add("write", on_dropdown_change)
    led_b_var.trace_add("write", on_dropdown_change)
    canvas_widget.bind("<Configure>", on_resize)
 
    process_queue()
    return root
 
async def run_ble_client(address):
    global running
    async with BleakClient(address) as client:
        try:
            print(f"Connected: {client.is_connected}")
            await client.start_notify(FSM_DATA_CHAR_UUID, notification_handler)
            print("FSM Data Monitoring Started. Close the plot window to stop.")
           
            while running:
                await asyncio.sleep(1)
       
        except Exception as e:
            print(f"Error: {e}")
        finally:
            try:
                await client.stop_notify(FSM_DATA_CHAR_UUID)
            except Exception as e:
                print(f"Error stopping notifications: {e}")
            print("FSM Data Monitoring Stopped.")
 
def run_asyncio_loop(loop):
    asyncio.set_event_loop(loop)
    loop.run_until_complete(run_ble_client(DEVICE_ADDRESS))
 
def save_to_excel(data, filename=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "FSM Data"
 
    # Write headers
    headers = ["System Time (s)", "Sample Time (s)"] + [f"Array {'A' if i < 21 else 'B'} {'W' + str((i%21)//3 + 1) if (i%21) < 18 else 'LED OFF'} PD{(i%21)%3 + 1}" for i in range(42)]
    ws.append(headers)
 
    # Write data
    for row in data:
        ws.append(row)
 
    # Set column widths and number format
    for col in range(1, 59):  # Increased by 1 to account for the new column
        ws.column_dimensions[get_column_letter(col)].width = 15
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col).number_format = '0.000000'
 
    # Generate filename with timestamp if not provided
    if filename is None:
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"fsm_data_{timestamp}.xlsx"
 
    # Save the workbook
    wb.save(filename)
    print(f"Data saved to {filename}")
    os.remove(filename_csv)

def csv_create():
    with open(filename_csv, 'w+') as file:
        headers = ["System Time (s)", "Sample Time (s)"] + [f"Array {'A' if i < 21 else 'B'} {'W' + str((i%21)//3 + 1) if (i%21) < 18 else 'LED OFF'} PD{(i%21)%3 + 1}" for i in range(42)]
        file.write(", ".join(headers))

def csv_update(timestamp, sample_time, array_a_data, array_b_data):
    return
    with open(filename_csv, 'a') as file:
        file.write("{}, {}".format(timestamp, sample_time))
        file.write(", ".join(map(str, array_a_data)))
        file.write(", ".join(map(str, array_b_data)))
        
def main():
    global running
    csv_create()
    loop = asyncio.new_event_loop()
    ble_thread = threading.Thread(target=run_asyncio_loop, args=(loop,))
    ble_thread.start()
 
    root = create_gui()
    root.mainloop()
 
    running = False
    ble_thread.join()
 
    collected_data = []
    for i in range(len(time_data)):
        row = [time_data[i], sample_time_data[i]] + [array_a_data[j][i] for j in range(21)] + [array_b_data[j][i] for j in range(21)]
        collected_data.append(row)
    save_to_excel(collected_data)
    print("Data saved to Excel file.")
 
if __name__ == "__main__":
    main()
